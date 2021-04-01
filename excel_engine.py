#!venv/Scripts/python
# -*- coding: utf-8 -*-

import openpyxl
import sqlite3
import sys
from tkinter import messagebox


def ordinal(n): return "%d%s" % (
    n, "tsnrhtdd"[(n//10 % 10 != 1)*(n % 10 < 4)*n % 10::4])


class ExcelReader:

    current_sheet = None

    sheet_headers = []

    conn = None

    def __init__(self, path_to_file):
        # define excel book filepath
        self.path_to_file = path_to_file
        # read excel book and set current working sheet
        self.readBook()
        # get titles from work sheet
        self.get_title_columns_name()
        # create temp.db and create connection to it
        self.sqlite_init()
        # create table in temp.db with column names as sheet_headers values
        self.create_data_table()
        # fill table with values from excel sheet
        self.fill_db_with_excel_data()
        # close connection to database
        self.conn.close()

    def readBook(self):
        wb = openpyxl.load_workbook(self.path_to_file)
        self.current_sheet = wb['Журнал активности']

    def sqlite_init(self):
        try:
            self.conn = sqlite3.connect("tmp.db")  # create temp database
            cursor = self.conn.cursor()  # create cursor object

            print("database has been successfully created and connected")

            __sqlite_select_query = "SELECT sqlite_version();"
            cursor.execute(__sqlite_select_query)
            record = cursor.fetchall()
            print(f"SQlite version: {record}")
        except sqlite3.Error as error:
            print(
                f"\033[91mAn error has occured when trying connect to sqlite\n {error}\033[0m")

    def get_title_columns_name(self):
        __valid_columns_name = (
            "IP адрес",
            "Количество символов в минуту",
            "Число запросов на сервер за 1 мин",
            "Число запросов на сервер за 5 мин",
            "Число неудачных попыток входа в систему за 5 мин"
        )
        for col in range(1, self.current_sheet.max_column + 1):
            if __valid_columns_name[col - 1] == self.current_sheet.cell(row=1, column=col).value:
                self.sheet_headers.append(self.current_sheet.cell(row=1, column=col).value)
            else:
                messagebox.showerror("Ошибка", "Неверный формат книги Excel")
                raise OSError

    def create_data_table(self):
        print("Starting to create main table...")
        try:
            __cursor = self.conn.cursor()
            __cursor.execute("""
                                CREATE TABLE IF NOT EXISTS activity_journal(
                                    id INT PRIMARY KEY,
                                    ip_address TEXT,
                                    symbols_per_min INT,
                                    requests_per_min INT,
                                    requests_per_5_min INT,
                                    unsuccessful_auths INT
                                )                          
                            """)
            self.conn.commit()
        except sqlite3.Error as error:
            print(
                f"\033[91mAn error has occured when tried create table 'activity_journal'\n{error}\033[0m")

    def fill_db_with_excel_data(self):
        __cursor = self.conn.cursor()
        cs = self.current_sheet
        tmp_row = 0
        errored_counter = 0
        try:
            for row in range(2, self.current_sheet.max_row + 1):
                tmp_row = row - 1

                if not __cursor.execute("SELECT * FROM activity_journal WHERE ip_address ='"+cs.cell(row=row, column=1).value + "'").fetchone():
                    try:
                        entry = (row - 1,
                                 cs.cell(row=row, column=1).value,
                                 int(cs.cell(row=row, column=2).value) or 0,
                                 int(cs.cell(row=row, column=3).value) or 0,
                                 int(cs.cell(row=row, column=4).value) or 0,
                                 int(cs.cell(row=row, column=5).value) or 0)
                        __cursor.execute(
                            "INSERT INTO activity_journal VALUES(?, ?, ?, ?, ?, ?)", entry)
                        self.conn.commit()
                    except ValueError as error:
                        errored_counter += 1
                        continue
                    except TypeError as t_error:
                        errored_counter += 1
                        continue

                sys.stdout.write(f"{round(row/self.current_sheet.max_row*100)} percent complete\r")
            if errored_counter == 0:
                print("\nData successfully imported")
            else:
                print(f"\nData imported\n\033[91m{errored_counter} entries were not imported due incompatible format\033[0m")
        except sqlite3.Error as error:
            messagebox.showerror("Ошибка", "Error has occured when tried to insert {ordinal(tmp_row)} entry\n{error}")
            print(
                f"\033[91mError has occured when tried to insert {ordinal(tmp_row)} entry\n{error}\033[0m")

def main():
    reader = ExcelReader("C:/Users/Postalet/Desktop/zhurnal_aktivnosti.xlsx")


if __name__ == "__main__":
    main()
